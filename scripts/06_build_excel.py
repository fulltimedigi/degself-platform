#!/usr/bin/env python3
"""Rebuild shuwaikh_workshops.xlsx from v15 JSON with same 4-sheet structure."""
import json
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('/home/user/workspace/shuwaikh_workshops_v19.json') as f:
    data = json.load(f)

print(f"Loaded {len(data)} workshops")

wb = openpyxl.Workbook()

# ---------- Sheet 1: Workshops ----------
ws = wb.active
ws.title = "Workshops"
ws.sheet_view.rightToLeft = True

headers = ['#', 'الاسم', 'المحافظة', 'المنطقة', 'التخصص', 'الفئة الأصلية',
           'العنوان', 'التلفون', 'الموقع', 'التقييم', 'المراجعات', 'الإحداثيات',
           'ساعات العمل', 'صورة', 'Google URL', 'place_id', 'المصدر']

header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='0A0A0A')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
body_font = Font(name='Calibri', size=10)
border = Border(left=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'),
                bottom=Side(style='thin', color='DDDDDD'))

for col_idx, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col_idx, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = header_align
    c.border = border

for i, w in enumerate(data, start=1):
    row = i + 1
    lat = w.get('latitude')
    lng = w.get('longitude')
    coords = f"{lat},{lng}" if lat is not None and lng is not None and lat != '' and lng != '' else ''

    values = [
        i,
        w.get('name', '') or '',
        w.get('governorate', '') or '',
        w.get('area', '') or '',
        w.get('specialty', '') or '',
        w.get('category_raw', '') or '',
        w.get('address', '') or '',
        w.get('phone', '') or '',
        w.get('website', '') or '',
        w.get('rating', '') if w.get('rating') is not None else '',
        w.get('reviews_count', '') if w.get('reviews_count') is not None else '',
        coords,
        w.get('opening_hours', '') or '',
        w.get('main_image', '') or '',
        w.get('google_url', '') or '',
        w.get('place_id', '') or '',
        w.get('source', '') or '',
    ]
    for col_idx, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col_idx, value=v)
        c.font = body_font
        c.border = border
        if col_idx in (1, 10, 11):
            c.alignment = Alignment(horizontal='center', vertical='center')
        else:
            c.alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)

# Column widths
widths = [5, 32, 16, 22, 20, 22, 40, 18, 28, 8, 10, 22, 38, 30, 32, 30, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 28
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

# ---------- Sheet 2: المحافظات ----------
ws2 = wb.create_sheet("المحافظات")
ws2.sheet_view.rightToLeft = True
gov_counts = Counter(w.get('governorate') or 'غير محدد' for w in data)
total = len(data)
ws2.append(['المحافظة', 'عدد الورش', 'النسبة'])
for c in ws2[1]:
    c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = border
for gov, n in sorted(gov_counts.items(), key=lambda x: -x[1]):
    ws2.append([gov, n, f"{n/total*100:.1f}%"])
for row in ws2.iter_rows(min_row=2):
    for c in row:
        c.font = body_font; c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')
ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 12

# ---------- Sheet 3: التخصصات ----------
ws3 = wb.create_sheet("التخصصات")
ws3.sheet_view.rightToLeft = True
spec_counts = Counter(w.get('specialty') or 'غير محدد' for w in data)
ws3.append(['التخصص', 'العدد', 'النسبة'])
for c in ws3[1]:
    c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = border
for spec, n in sorted(spec_counts.items(), key=lambda x: -x[1]):
    ws3.append([spec, n, f"{n/total*100:.1f}%"])
for row in ws3.iter_rows(min_row=2):
    for c in row:
        c.font = body_font; c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')
ws3.column_dimensions['A'].width = 28
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 12

# ---------- Sheet 4: جودة البيانات ----------
ws4 = wb.create_sheet("جودة البيانات")
ws4.sheet_view.rightToLeft = True

def has(field):
    return sum(1 for w in data if w.get(field) not in (None, '', [], {}))
def has_coords():
    return sum(1 for w in data if w.get('latitude') not in (None, '') and w.get('longitude') not in (None, ''))

metrics = [
    ('إجمالي الورش', total),
    ('لها تلفون', has('phone')),
    ('لها عنوان', has('address')),
    ('لها إحداثيات GPS', has_coords()),
    ('لها ساعات عمل', has('opening_hours')),
    ('لها صورة رئيسية', has('main_image')),
    ('لها تقييم', has('rating')),
    ('لها Google URL', has('google_url')),
    ('لها محافظة', has('governorate')),
]
ws4.append(['المعيار', 'العدد', 'النسبة'])
for c in ws4[1]:
    c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = border
for label, n in metrics:
    ws4.append([label, n, f"{n/total*100:.1f}%"])
for row in ws4.iter_rows(min_row=2):
    for c in row:
        c.font = body_font; c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')
ws4.column_dimensions['A'].width = 26
ws4.column_dimensions['B'].width = 12
ws4.column_dimensions['C'].width = 12

wb.save('/home/user/workspace/shuwaikh_workshops.xlsx')
print("Saved: /home/user/workspace/shuwaikh_workshops.xlsx")
print("\nData quality (v15):")
for label, n in metrics:
    print(f"  {label:25s}: {n:5d} ({n/total*100:.1f}%)")
